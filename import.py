# -*- coding: utf-8 -*-
"""
Created on Mon Jun 21 10:50:11 2021

@author: jeanneau
"""
        
## importation des modules

import os
from openpyxl import load_workbook
import pandas as pd
import re
from datetime import datetime
import pymongo  
import json
import locale
locale.setlocale(locale.LC_TIME,'')
import math

## creation d'une fonction pour permettre de diviser par 0

def divi(x, y):
    return 0 if y == 0 else x / y


########################## 2017 ##################################################################################

## choix du chemin d'accès du dossier où se trouvent les .xlsx de 2017

directory = os.fsencode('C:\\Users\\jeanneau\\Desktop\\sirac\\2017')

## boucle pour récupérer les noms de tous les .xlsx dans le dossier spécifié juste au dessus

files = []
    
for file in os.listdir(directory):
     filename = os.fsdecode(file)
     files.append(filename)

## spécifier encore une fois le chemin mais cette fois en format string suivi de \\

path = 'C:\\Users\\jeanneau\\Desktop\\sirac\\2017\\'

## Début de la boucle qui "nettoie" chacun des .xslx et qui les insert dans MongoDB

for x in files:

    ### chargement du .xslx et conversion en dataframe    

    wb = load_workbook(path+x)
    ws = wb.active
    df = pd.DataFrame(ws.values)
    
    ## récupération des dates observés dans une liste
      
    j=4
    list_of_day=[]

    while j<len(df):
        jour=df.at[j,7]
        j+=32
        list_of_day.append(jour)
    
    ## convertion des dates en format unix
    
    unix_list = []
    for a in range(len(list_of_day)):
        unix = datetime.strptime(list_of_day[a], "%A %d %B %Y")
        timestamp = datetime.timestamp(unix)
        unix_list.append(int(timestamp))
        
    ## récupération du jour seulement (lundi, mardi,...)
        
    day_list = []
    for a in range(len(list_of_day)):
        day =  re.findall(r'^(.*?) ', list_of_day[a])
        day_list.append(day[0])
    
    ## récupération du chiffre de la date (1,2,...,31)
        
    number_list = []
    for a in range(len(list_of_day)):
        number =  re.findall(r' (.*?) ', list_of_day[a])
        number_list.append(int(number[0]))
    
    ## récupération du mois 
    
    month_list = []
    for a in range(len(list_of_day)):
        month =  re.findall(r'[0-9] (.*?) [0-9]', list_of_day[a])
        month_list.append(month[0])
        
    ## récupération de l'année
        
    year_list = []
    for a in range(len(list_of_day)):
        year =  re.findall(r'[0-9]{4}$', list_of_day[a])
        year_list.append(int(year[0]))

    ## nettoyage de la cellule qui contient le nom du trajet (retirer les espaces, la longueur,...)
    
    trajet=df.at[3,7]
    trajet = re.sub(r"\s+", "", trajet)
    trajet = re.sub(r">", "", trajet)
    trajet = re.sub(r"\.","", trajet)
    trajet = re.sub(r"\:","", trajet)
    trajet = re.sub(r'Longueur(.*?)m',"",trajet)
    
    ## récupération de la longueur:
    
    j=3
    list_of_long=[]

    while j<len(df):
        m=df.at[j,7]
        long= re.findall(r'Longueur:(.*?)m',m )
        j+=32
        list_of_long.append(int(long[0]))

    
    ## suppression des collones vides
    
    columns_to_keep = [x for x in range(df.shape[1]) if x not in [0,1,6,7,10,12,21,22,24,29,30,31,32,33,34,35]]
    data = df.iloc[:, columns_to_keep]

    ## Récupération des temps de parcours pour chaque jour (à intervalle régulier dans le fichier)

    m=7
    n=31
    list_of_data=[]

    while m < len(data):
        data_ = data[m:n]
        m+=32
        n+=32
        list_of_data.append(data_)
     
    ## convertion des temps de parcours en secondes (en passant par la séparation en parties entières et décimales) puis en km/h
    
    dataList = []
           
    for i in range(len(list_of_data)):
       df = list_of_data[i]
        
       for row in df.iterrows(): 
           mylist = row[1].tolist()
           mylist = [i.replace(':','.') for i in mylist]
           mylist= [float(j) for j in mylist]
           mylist = [math.modf(y) for y in mylist]
           mylist= [int((x[1]*60)+(x[0]*100)) for x in mylist]
           mylist= [(divi(int(long[0]),x))*3.6 for x in mylist]
           mylist = [x if x<90 else 90 for x in mylist]

           
           dataList.append(mylist)
           
     
    ## connexion au serveur MongoDB, création de la database ADEUS, et de la collection qui porte le nom du trajet
    
    client= pymongo.MongoClient('localhost', 27017)
    mydb= client["ADEUS"]
    collection= mydb[trajet]

    ## création du dictionnaire qui contient toutes les variables précédemment crées (jour, nombre, mois année, unix et temps de parcours)
         
    dictionary = []
        
    for a in range(len(unix_list)):
        listofTuples = [ ("day",str(day_list[a])),("number",number_list[a]),("month", str(month_list[a])), ("year", year_list[a]), ("unix", unix_list[a]),("longeur",list_of_long[a]),("speed",dataList[a])]
        dictionary.append(dict(listofTuples))
    
         
    ## remplacement des vitesses dans le dictionnaire par les vitesses correctes (heure par heure)
    
    a=0
    b=0
    c=24
    while c <= len(dataList):       
        dictionary[a].update(dict(speed=dataList[b:c]))
        a+=1
        b+=24
        c+=24

    ## insertition du dictionnaire dans la collection qui porte le nom du trajet 
           
    collection.insert_many(dictionary)
    
########################## 2018/2019 ##################################################################################

## choix du chemin d'accès du dossier où se trouvent les .xlsx de 2017

directory = os.fsencode('C:\\Users\\jeanneau\\Desktop\\sirac\\20182019')

## boucle pour récupérer les noms de tous les .xlsx dans le dossier spécifié juste au dessus

files = []
    
for file in os.listdir(directory):
     filename = os.fsdecode(file)
     files.append(filename)

## spécifier encore une fois le chemin mais cette fois en format string suivi de \\

path = 'C:\\Users\\jeanneau\\Desktop\\sirac\\20182019\\'

## Début de la boucle qui "nettoie" chacun des .xslx et qui les insert dans MongoDB

for x in files:

    ### chargement du .xslx et conversion en dataframe    

    wb = load_workbook(path+x)
    ws = wb.active
    df = pd.DataFrame(ws.values)
    
    ## récupération des dates observés dans une liste
      
    j=4
    list_of_day=[]

    while j<len(df):
        jour=df.at[j,7]
        j+=32
        list_of_day.append(jour)
    
    ## convertion des dates en format unix
    
    unix_list = []
    for a in range(len(list_of_day)):
        unix = datetime.strptime(list_of_day[a], "%A %d %B %Y")
        timestamp = datetime.timestamp(unix)
        unix_list.append(int(timestamp))
        
    ## récupération du jour seulement (lundi, mardi,...)
        
    day_list = []
    for a in range(len(list_of_day)):
        day =  re.findall(r'^(.*?) ', list_of_day[a])
        day_list.append(day[0])
    
    ## récupération du chiffre de la date (1,2,...,31)
        
    number_list = []
    for a in range(len(list_of_day)):
        number =  re.findall(r' (.*?) ', list_of_day[a])
        number_list.append(int(number[0]))
    
    ## récupération du mois 
    
    month_list = []
    for a in range(len(list_of_day)):
        month =  re.findall(r'[0-9] (.*?) [0-9]', list_of_day[a])
        month_list.append(month[0])
        
    ## récupération de l'année
        
    year_list = []
    for a in range(len(list_of_day)):
        year =  re.findall(r'[0-9]{4}$', list_of_day[a])
        year_list.append(int(year[0]))

    ## nettoyage de la cellule qui contient le nom du trajet (retirer les espaces, la longueur,...)
    
    trajet=df.at[3,7]
    trajet = re.sub(r"\s+", "", trajet)
    trajet = re.sub(r">", "", trajet)
    trajet = re.sub(r"\.","", trajet)
    trajet = re.sub(r"\:","", trajet)
    trajet = re.sub(r'Longueur(.*?)m',"",trajet)
    
    ## récupération de la longueur:
    
    j=3
    list_of_long=[]

    while j<len(df):
        m=df.at[j,7]
        long= re.findall(r'Longueur:(.*?)m',m )
        j+=32
        list_of_long.append(long)

    
    ## suppression des collones vides
    
    columns_to_keep = [x for x in range(df.shape[1]) if x not in [0,1,6,7,19,20,22,27,28,29,30,31]]
    data = df.iloc[:, columns_to_keep]

    ## Récupération des temps de parcours pour chaque jour (à intervalle régulier dans le fichier)

    m=7
    n=31
    list_of_data=[]

    while m < len(data):
        data_ = data[m:n]
        m+=32
        n+=32
        list_of_data.append(data_)
        
           ## convertion des temps de parcours en secondes (en passant par la séparation en parties entières et décimales)    
    
    dataList = []
           
    for i in range(len(list_of_data)):
       df = list_of_data[i]
        
       for row in df.iterrows(): 
           mylist = row[1].tolist()
           mylist = [i.replace(':','.') for i in mylist]
           mylist= [float(j) for j in mylist]
           mylist = [math.modf(y) for y in mylist]
           mylist= [int((x[1]*60)+(x[0]*100)) for x in mylist]
           mylist= [(divi(int(long[0]),x))*3.6 for x in mylist]
           mylist = [x if x<90 else 90 for x in mylist]


           
           dataList.append(mylist)
           
     
    ## connexion au serveur MongoDB, création de la database ADEUS, et de la collection qui porte le nom du trajet
    
    client= pymongo.MongoClient('localhost', 27017)
    mydb= client["ADEUS"]
    collection= mydb[trajet]

    ## création du dictionnaire qui contient toutes les variables précédemment crées (jour, nombre, mois année, unix et temps de parcours)
         
    dictionary = []
        
    for a in range(len(unix_list)):
        listofTuples = [ ("day",str(day_list[a])),("number",number_list[a]),("month", str(month_list[a])), ("year", year_list[a]), ("unix", unix_list[a]),("longeur",int(list_of_long[a][0])),("speed",dataList[a])]
        dictionary.append(dict(listofTuples))
    
 
    ## remplacement des vitesses dans le dictionnaire par les vitesses correctes (heure par heure)
         
    a=0
    b=0
    c=24
    while c <= len(dataList):       
        dictionary[a].update(dict(speed=dataList[b:c]))
        a+=1
        b+=24
        c+=24

    ## insertition du dictionnaire dans la collection qui porte le nom du trajet 
           
    collection.insert_many(dictionary)
